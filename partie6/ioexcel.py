import openpyxl

from pathlib import Path
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from utilitaires.data_utils import load_trajets_dict

pa = Path("../data") / "systeme.json"

def essai_write_excel():
    trajets = load_trajets_dict(pa)

    # Création du fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Trajets"

    # Écriture des en-têtes
    ws.append(["user", "depart", "arrivee", "distance"])

    # Écriture des trajets
    for trajet in trajets:
        ws.append([trajet["user"], trajet["depart"], trajet["arrivee"], trajet["distance"]])

    wb.save("trajets.xlsx")

def essai_read_excel():
    wb = load_workbook("trajets.xlsx")
    ws = wb["Trajets"]

    # Récupération des en-têtes
    headers = [cell.value for cell in ws[1]]  # Ligne 1

    # Lecture des données en tant que dictionnaires
    liste_trajets = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        trajet = dict(zip(headers, row))
        liste_trajets.append(trajet)

    # Affichage
    for t in liste_trajets:
        print(t)
